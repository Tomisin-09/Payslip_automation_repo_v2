from openpyxl import load_workbook
import logging


def write_employee_to_template(
    template_path: str,
    output_path: str,
    employee_ref: str,
    employee_ref_named_range: str,
):
    logger = logging.getLogger(__name__)

    logger.info(
        "Generating payslip for employee ref %s",
        employee_ref
    )

    # --------------------------------------------------
    # Load workbook
    # --------------------------------------------------
    wb = load_workbook(template_path, data_only=False)

    # --------------------------------------------------
    # Resolve named range
    # --------------------------------------------------
    if employee_ref_named_range not in wb.defined_names:
        raise ValueError(
            f"Named range '{employee_ref_named_range}' not found in workbook"
        )

    defined_range = wb.defined_names[employee_ref_named_range]
    dests = list(defined_range.destinations)

    if not dests:
        raise ValueError(
            f"Named range '{employee_ref_named_range}' has no destinations"
        )

    sheet_name, cell_ref = dests[0]
    ws = wb[sheet_name]

    # --------------------------------------------------
    # Set employee reference (THIS IS THE KEY STEP)
    # --------------------------------------------------
    ws[cell_ref] = employee_ref

    logger.info(
        "Set %s = %s on sheet '%s'",
        employee_ref_named_range,
        employee_ref,
        sheet_name,
    )

    # --------------------------------------------------
    # Save output workbook
    # --------------------------------------------------
    wb.save(output_path)

    logger.info("Saved XLSX: %s", output_path)
